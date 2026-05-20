"""
Rex Airlines — PER ↔ MJK Scraper (Playwright + Bright Data Scraping Browser)
=============================================================================
Only 2 routes: PER → MJK  and  MJK → PER

Migration from Selenium → Playwright:
  - Uses Bright Data Scraping Browser via CDP (wss:// port 9222)
  - CDP gives direct reCAPTCHA solve events — no more blind polling
  - Playwright auto-waits for elements — fewer manual time.sleep() needed
  - Network interception to detect flight result API calls early
  - Same Excel output format as original Selenium version

Flow for each date:
  1. Connect to Bright Data Scraping Browser (fresh context per date)
  2. Open Rex homepage
  3. Select one-way trip
  4. Fill origin / destination
  5. Set departure date
  6. Submit form — wait for CAPTCHA + verification
  7. Extract flight data
  8. Save to Excel
  9. Close context → move to next date

Install:
  pip install playwright openpyxl
  playwright install chromium
"""

from __future__ import annotations

import os
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
import re
import sys
import json
import time
import random
import tempfile
import traceback
import argparse
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo  # type: ignore[no-redef]

from playwright.sync_api import (
    sync_playwright, Page, Browser, BrowserContext,
    TimeoutError as PWTimeout, Error as PWError
)
from openpyxl import load_workbook, Workbook

try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass  # Already wrapped or read-only on some server envs — safe to ignore


# ─────────────────────────────────────────────────────────────
#  BRIGHT DATA CREDENTIALS
#  Scraping Browser uses port 9222 (CDP/WebSocket) — NOT 9515
# ─────────────────────────────────────────────────────────────
BD_BROWSER_HOST          = os.getenv("BD_BROWSER_HOST",    "brd.superproxy.io")
BD_SCRAPING_BROWSER_PORT = os.getenv("BD_SCRAPING_BROWSER_PORT", "9222")
BD_BROWSER_USER          = os.getenv("BD_BROWSER_USER",    "brd-customer-hl_fbc4a16a-zone-cont_rex")
BD_BROWSER_PASS          = os.getenv("BD_BROWSER_PASS",    "072res2p22t3")
BD_AUTH_TOKEN            = os.getenv("BD_AUTH_TOKEN",      "7b1cdf1c-e4e0-4b6c-925b-0121031e6bf7")
BD_WEB_UNLOCKER_ZONE     = os.getenv("BD_WEB_UNLOCKER_ZONE", "cron_rex")
BD_UNLOCKER_COUNTRY      = os.getenv("BD_UNLOCKER_COUNTRY", "au")

# WebSocket CDP endpoint for Playwright
BD_CDP_URL = (
    f"wss://{BD_BROWSER_USER}:{BD_BROWSER_PASS}"
    f"@{BD_BROWSER_HOST}:{BD_SCRAPING_BROWSER_PORT}"
)

# ─────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────
ROUTES = [("PER", "MJK"), ("MJK", "PER")]

AIRPORT_MAP = {
    "PER": "Perth",
    "MJK": "Monkey Mia (Shark Bay)",
}

REX_TIMEZONE  = os.getenv("REX_TIMEZONE", "Australia/Perth")
REX_TZ        = ZoneInfo(REX_TIMEZONE)
TOTAL_DAYS    = int(os.getenv("REX_TOTAL_DAYS", "84"))
START_OFFSET  = int(os.getenv("REX_START_OFFSET_DAYS", "0"))
_RUN_TS       = datetime.now(ZoneInfo(os.getenv("REX_TIMEZONE", "Australia/Perth"))).strftime("%Y%m%d_%H%M%S")
OUTPUT_EXCEL  = os.getenv("REX_OUTPUT_EXCEL", f"rex_per_mjk_results_{_RUN_TS}.xlsx")
DEBUG_DIR     = os.getenv("REX_DEBUG_DIR", "rex_debug")
LOG_DIR       = os.getenv("REX_LOG_DIR", "rex_logs")
RUN_ID        = os.getenv("REX_RUN_ID", datetime.now(REX_TZ).strftime("%Y%m%d"))
MAX_ATTEMPTS  = int(os.getenv("REX_MAX_ATTEMPTS", "3"))
PAGE_TIMEOUT  = int(os.getenv("REX_PAGE_TIMEOUT", "180"))   # seconds
CAPTCHA_WAIT  = int(os.getenv("REX_CAPTCHA_WAIT", "90"))    # seconds

STATUS_SUCCESS  = "SUCCESS"
STATUS_NO_FARE  = "NO_FARE_AVAILABLE"
STATUS_FAILED   = "FAILED_AFTER_RETRIES"
STATUS_TIMEOUT  = "PAGE_TIMEOUT"
STATUS_BLOCKED  = "BLOCKED_OR_VERIFICATION_REQUIRED"
STATUS_STRUCT   = "POSSIBLE_WEBSITE_STRUCTURE_ISSUE"

CSV_FIELDS = [
    "Date Checked", "Time Checked", "Airline",
    "Date of Departure", "Time of Departure",
    "Origin", "Destination",
    "Fare Price", "Fare Class", "Source",
    "Run ID", "Status", "Comment", "Retry Count", "Debug Artifacts",
]

COMPLETED_STATUSES = {STATUS_SUCCESS, STATUS_NO_FARE}
RESUME_MODE        = False


# ─────────────────────────────────────────────────────────────
#  UTILITIES
# ─────────────────────────────────────────────────────────────

def send_email_with_attachment(subject: str, body: str, file_path: Optional[str] = None) -> bool:
    smtp_email = os.getenv("SMTP_EMAIL") or os.getenv("EMAIL_FROM")
    smtp_password = os.getenv("SMTP_PASSWORD") or os.getenv("EMAIL_PASSWORD")
    recipient_email = os.getenv("RECIPIENT_EMAIL") or os.getenv("EMAIL_TO")
    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port_str = os.getenv("SMTP_PORT", "587")

    if not smtp_email or not smtp_password or not recipient_email:
        print("⚠️  EMAIL_FROM, EMAIL_PASSWORD, or EMAIL_TO not set in environment. Skipping email report.")
        return False

    try:
        smtp_port = int(smtp_port_str)
    except ValueError:
        smtp_port = 587

    print(f"📧 Sending email report to {recipient_email} via SMTP...")
    try:
        msg = MIMEMultipart()
        msg["From"] = smtp_email
        msg["To"] = recipient_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        if file_path and os.path.exists(file_path):
            filename = os.path.basename(file_path)
            with open(file_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {filename}",
            )
            msg.attach(part)
            print(f"📎 Attached file: {filename}")

        # Connect to SMTP server
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_email, smtp_password)
        server.send_message(msg)
        server.quit()
        print("✅ Email report sent successfully!")
        return True
    except Exception as e:
        print(f"❌ Failed to send email: {e}")
        traceback.print_exc()
        return False


def rex_now() -> datetime:
    return datetime.now(REX_TZ).replace(tzinfo=None)


def today_dt() -> datetime:
    return rex_now().replace(hour=0, minute=0, second=0, microsecond=0)


def build_date_list() -> list[datetime]:
    t = today_dt() + timedelta(days=START_OFFSET)
    return [t + timedelta(days=i) for i in range(TOTAL_DAYS)]


def output_date(dt: datetime) -> str:
    return dt.strftime("%d-%m-%Y")


def rex_form_date(dt: datetime) -> str:
    """Rex form expects '20 May 2026' format."""
    return dt.strftime("%d %b %Y")


def _format_price(dollars: str, cents: Optional[str] = None) -> str:
    dollars = re.sub(r"\s+", "", dollars).replace(",", "").strip()
    if cents is None:
        return f"${dollars}.00" if "." not in dollars else f"${dollars}"
    return f"${dollars}.{cents.strip()}"


def _extract_price_patterns(text: str) -> str:
    """
    Rex renders prices in fragmented ways — normalise and try all patterns.

    Rex shows multiple fare tiers (Community, Flex, etc.) on the same page.
    We always want the LOWEST price (Community/Economy fare), NOT Flex.

    Strategy:
      1. Try to find the Community section first and extract price from there.
      2. If not found, collect ALL prices in the window and return the lowest.
    """
    flat = re.sub(r"[\s\n\r]+", " ", text)

    # ── Priority 1: Community section price ──────────────────────────────────
    # Rex page structure: "Community ... $264.65 ... Flex ... $775.21"
    comm_pos = flat.lower().find("community")
    flex_pos = flat.lower().find("flex")

    # Determine the Community section boundaries
    if comm_pos >= 0:
        # End of Community section = start of Flex (or 400 chars, whichever first)
        comm_end = flex_pos if (flex_pos > comm_pos) else (comm_pos + 400)
        comm_section = flat[comm_pos:comm_end]
        patterns = [
            r"[Ff]rom\s*\$\s*([\d,]+)\.(\d{2})",
            r"\$\s*([\d,]+)\.(\d{2})",
            r"[Ff]rom\s*\$\s*([\d,]+)\s+(\d{2})(?!\d)",
            r"\$\s*([\d,]+)\s+(\d{2})(?!\d)",
        ]
        for pattern in patterns:
            m = re.search(pattern, comm_section)
            if m:
                dollars_raw = re.sub(r"\s+", "", m.group(1))
                return _format_price(dollars_raw, m.group(2))

    # ── Priority 2: Collect ALL prices and return the lowest ─────────────────
    # This guards against page layout changes — always picks cheapest fare.
    all_price_pattern = r"\$\s*([\d,]+)\.(\d{2})"
    all_matches = re.findall(all_price_pattern, flat)
    if all_matches:
        prices = []
        for dollars_raw, cents in all_matches:
            dollars_clean = re.sub(r"[\s,]+", "", dollars_raw)
            try:
                val = float(f"{dollars_clean}.{cents}")
                prices.append((val, dollars_clean, cents))
            except ValueError:
                continue
        if prices:
            prices.sort(key=lambda x: x[0])   # ascending — lowest first
            _, d, c = prices[0]
            return _format_price(d, c)

    # ── Fallback: fragmented / space-separated cents ─────────────────────────
    patterns_fallback = [
        r"[Ff]rom\s*\$\s*([\d,]+)\s+(\d{2})(?!\d)",
        r"\$\s*([\d,]+)\s+(\d{2})(?!\d)",
        r"[Ff]rom\s*\$\s*(\d[\d,\s]*\d)\s*[\.\s]\s*(\d{2})(?!\d)",
        r"([\d,]+)\s*\.\s*(\d{2})\s+[Ss]elect\s+[Ff]ares",
    ]
    for pattern in patterns_fallback:
        m = re.search(pattern, flat)
        if m:
            dollars_raw = re.sub(r"\s+", "", m.group(1))
            return _format_price(dollars_raw, m.group(2))

    sf_pos = flat.lower().find("select fares")
    if sf_pos > 0:
        before = flat[:sf_pos]
        m = re.search(r"(\d[\d,]*)\s*[.\s]\s*(\d{2})\s*$", before.strip())
        if m:
            return _format_price(re.sub(r"\s+", "", m.group(1)), m.group(2))
        m = re.search(r"\$\s*(\d[\d,]*)", before)
        if m:
            return _format_price(m.group(1))

    m = re.search(r"[Ff]rom\s*\$\s*(\d[\d,]*)", flat)
    if m:
        return _format_price(m.group(1))
    m = re.search(r"\$\s*(\d[\d,]*)", flat)
    if m:
        return _format_price(m.group(1))
    return "N/A"


def extract_all_times(text: str) -> list[str]:
    times = re.findall(r"\d{1,2}:\d{2}\s*[aApP][mM]", text)
    if times:
        return [t.strip() for t in times]
    return [t.strip() for t in re.findall(r"\b\d{1,2}:\d{2}\b", text)]


def extract_flights_from_body(body_text: str, origin: str, dest: str, date_str: str) -> list[dict]:
    """
    Extract ZL flight numbers and prices from page body text.
    PER→MJK is a direct non-stop flight — standard price extraction applies.
    """
    now = rex_now()
    ck_date = now.strftime("%d-%m-%Y")
    ck_time = now.strftime("%H:%M:%S")

    markers = [
        "Select your departing flight", "Departure Time",
        "departing flight", "Fly Economy", "Select Fares",
    ]
    ribbon_end = len(body_text) // 7
    for marker in markers:
        pos = body_text.find(marker)
        if pos > 0:
            ribbon_end = pos
            break
    zl_match = re.search(r"ZL\s?\d{3,4}", body_text)
    if zl_match and zl_match.start() > 100:
        ribbon_end = min(ribbon_end, max(0, zl_match.start() - 50))

    print(f"   📍 Ribbon area ends at ~char {ribbon_end}")
    body_after = body_text[ribbon_end:]
    all_times  = extract_all_times(body_after)
    zl_matches = list(re.finditer(r"ZL\s?\d{3,4}", body_after))

    print(f"   🔍 ZL matches after ribbon: {[m.group() for m in zl_matches]}")

    flights = []
    seen    = set()

    for idx, m in enumerate(zl_matches):
        f_no = re.sub(r"\s", "", m.group())

        dep_idx = idx * 2
        if dep_idx < len(all_times):
            dep = all_times[dep_idx]
        elif idx < len(all_times):
            dep = all_times[idx]
        else:
            dep = "-"

        after_start = m.end()
        after_end   = min(len(body_after), after_start + 900)
        window      = body_after[after_start:after_end]
        price       = _extract_price_patterns(window)

        key = f"{f_no}-{dep}"
        if key in seen:
            continue
        seen.add(key)

        print(f"      ✈️  {f_no}  dep={dep}  price={price}")
        flights.append({
            "Date Checked":      ck_date,
            "Time Checked":      ck_time,
            "Airline":           f_no,
            "Date of Departure": date_str,
            "Time of Departure": dep,
            "Origin":            origin,
            "Destination":       dest,
            "Fare Price":        price,
            "Fare Class":        "Economy",
            "Source":            "Rex Website",
            "Run ID":            RUN_ID,
            "Status":            STATUS_SUCCESS,
            "Comment":           "Flight and fare data parsed from Rex availability page",
            "Retry Count":       0,
            "Debug Artifacts":   "",
        })

    return flights


def make_no_fare_row(date_str: str, origin: str, dest: str, comment: str, retry_count: int = 0) -> dict:
    now = rex_now()
    return {
        "Date Checked":      now.strftime("%d-%m-%Y"),
        "Time Checked":      now.strftime("%H:%M:%S"),
        "Airline":           "no flight",
        "Date of Departure": date_str,
        "Time of Departure": "-",
        "Origin":            origin,
        "Destination":       dest,
        "Fare Price":        "N/A",
        "Fare Class":        "",
        "Source":            "Rex Website",
        "Run ID":            RUN_ID,
        "Status":            STATUS_NO_FARE,
        "Comment":           comment,
        "Retry Count":       retry_count,
        "Debug Artifacts":   "",
    }


def make_failed_row(date_str: str, origin: str, dest: str, status: str,
                    comment: str, retry_count: int, debug: str = "") -> dict:
    now = rex_now()
    return {
        "Date Checked":      now.strftime("%d-%m-%Y"),
        "Time Checked":      now.strftime("%H:%M:%S"),
        "Airline":           "scrape failed",
        "Date of Departure": date_str,
        "Time of Departure": "-",
        "Origin":            origin,
        "Destination":       dest,
        "Fare Price":        "N/A",
        "Fare Class":        "",
        "Source":            "Rex Website - failed",
        "Run ID":            RUN_ID,
        "Status":            status,
        "Comment":           comment,
        "Retry Count":       retry_count,
        "Debug Artifacts":   debug,
    }


# ─────────────────────────────────────────────────────────────
#  EXCEL OUTPUT  (unchanged from Selenium version)
# ─────────────────────────────────────────────────────────────

class OutputStore:
    def __init__(self, path: str):
        self.path = path

    def _load(self):
        if os.path.exists(self.path):
            wb = load_workbook(self.path)
            ws = wb.active
            first_has_values = any(
                ws.cell(row=1, column=c).value
                for c in range(1, len(CSV_FIELDS) + 1)
            )
            if not first_has_values:
                ws.append(CSV_FIELDS)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Rex PER MJK Data"
            ws.append(CSV_FIELDS)
        return wb, ws

    def _headers(self, ws) -> list[str]:
        return [cell.value or "" for cell in ws[1]]

    def _save(self, wb):
        out_dir = os.path.dirname(os.path.abspath(self.path)) or "."
        os.makedirs(out_dir, exist_ok=True)
        fd, tmp = tempfile.mkstemp(
            prefix=f".{Path(self.path).stem}_", suffix=".xlsx", dir=out_dir
        )
        os.close(fd)
        try:
            wb.save(tmp)
            os.replace(tmp, self.path)
        finally:
            if os.path.exists(tmp):
                try:
                    os.remove(tmp)
                except OSError:
                    pass

    def _row_matches(self, ws, row_idx, origin, dest, date_str) -> bool:
        headers = self._headers(ws)
        vals = {h: ws.cell(row=row_idx, column=i + 1).value for i, h in enumerate(headers)}
        return (
            str(vals.get("Run ID") or "") == RUN_ID
            and str(vals.get("Origin") or "") == origin
            and str(vals.get("Destination") or "") == dest
            and str(vals.get("Date of Departure") or "") == date_str
        )

    def write_rows(self, origin: str, dest: str, date_str: str, rows: list[dict]):
        wb, ws = self._load()
        headers = self._headers(ws)
        for row_idx in range(ws.max_row, 1, -1):
            if self._row_matches(ws, row_idx, origin, dest, date_str):
                ws.delete_rows(row_idx, 1)
        for row in rows:
            ws.append([row.get(f, "") for f in headers])
        self._save(wb)
        print(f"   💾 Saved {len(rows)} row(s) → {self.path}")

    def job_completed(self, origin: str, dest: str, date_str: str) -> bool:
        if not os.path.exists(self.path):
            return False
        wb, ws = self._load()
        headers = self._headers(ws)
        for row_idx in range(2, ws.max_row + 1):
            if self._row_matches(ws, row_idx, origin, dest, date_str):
                vals = {h: ws.cell(row=row_idx, column=i + 1).value for i, h in enumerate(headers)}
                if str(vals.get("Status") or "") in COMPLETED_STATUSES:
                    return True
        return False


# ─────────────────────────────────────────────────────────────
#  DEBUG ARTIFACTS  (Playwright version)
# ─────────────────────────────────────────────────────────────

def save_debug(page: Page, label: str, meta: Optional[dict] = None) -> str:
    """Save screenshot + page HTML + JSON metadata for failed jobs."""
    os.makedirs(DEBUG_DIR, exist_ok=True)
    safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", label).strip("_") or "page"
    ts = rex_now().strftime("%Y%m%d_%H%M%S")
    artifacts = []

    png_path = os.path.join(DEBUG_DIR, f"{safe}_{ts}.png")
    try:
        page.screenshot(path=png_path, full_page=True)
        artifacts.append(os.path.abspath(png_path))
        print(f"   📸 Screenshot: {png_path}")
    except Exception as e:
        print(f"   ⚠️  Screenshot failed: {e}")

    html_path = os.path.join(DEBUG_DIR, f"{safe}_{ts}.html")
    try:
        with open(html_path, "w", encoding="utf-8") as fh:
            fh.write(page.content())
        artifacts.append(os.path.abspath(html_path))
        print(f"   🧾 HTML: {html_path}")
    except Exception as e:
        print(f"   ⚠️  HTML dump failed: {e}")

    json_path = os.path.join(DEBUG_DIR, f"{safe}_{ts}.json")
    try:
        m = dict(meta or {})
        m.setdefault("url", page.url)
        m.setdefault("saved_at", rex_now().isoformat(sep=" "))
        try:
            body = page.inner_text("body")
            m["body_snippet"] = body[:1500]
        except Exception:
            pass
        with open(json_path, "w", encoding="utf-8") as fh:
            json.dump(m, fh, ensure_ascii=False, indent=2)
        artifacts.append(os.path.abspath(json_path))
        print(f"   🧭 Metadata: {json_path}")
    except Exception as e:
        print(f"   ⚠️  Metadata dump failed: {e}")

    return " | ".join(artifacts)


# ─────────────────────────────────────────────────────────────
#  PLAYWRIGHT BROWSER FACTORY
# ─────────────────────────────────────────────────────────────

def make_page(pw_instance, max_attempts: int = 4) -> tuple[Browser, BrowserContext, Page]:
    """
    Connect to Bright Data Scraping Browser via CDP (WebSocket).
    Returns (browser, context, page) — all three need to be closed on cleanup.

    Bright Data Scraping Browser:
      - Endpoint: wss://user:pass@brd.superproxy.io:9222
      - Handles reCAPTCHA, fingerprinting, IP rotation automatically
      - CDP protocol gives us direct event access (unlike Selenium)
    """
    last_exc = None
    for attempt in range(1, max_attempts + 1):
        browser = None
        try:
            print(f"   🔌 Connecting to Bright Data Scraping Browser (attempt {attempt}/{max_attempts})...")
            browser = pw_instance.chromium.connect_over_cdp(
                BD_CDP_URL,
                timeout=60_000,   # 60s connection timeout
            )
            # NOTE: Bright Data Scraping Browser manages its own headers, locale,
            # and timezone — do NOT pass extra_http_headers, locale, or timezone_id;
            # those raise "Overriding Accept-Language headers forbidden" errors.
            context = browser.new_context(
                viewport={"width": 1366, "height": 900},
            )
            context.set_default_timeout(30_000)           # 30s for element waits
            context.set_default_navigation_timeout(max(PAGE_TIMEOUT * 1000, 180_000))

            page = context.new_page()
            print("   ✅ Browser connected")
            return browser, context, page

        except Exception as exc:
            last_exc = exc
            msg = str(exc).lower()
            is_transient = any(x in msg for x in [
                "internal server error", "500", "connection",
                "timeout", "target closed", "websocket"
            ])
            if browser:
                try:
                    browser.close()
                except Exception:
                    pass
            if is_transient and attempt < max_attempts:
                wait = 8 * attempt
                print(f"   ⚠️  Bright Data connect error, retrying in {wait}s: {exc}")
                time.sleep(wait)
                continue
            raise


# ─────────────────────────────────────────────────────────────
#  PAGE STATE HELPERS
# ─────────────────────────────────────────────────────────────

def page_body_text(page: Page) -> str:
    try:
        return page.inner_text("body")
    except Exception:
        return ""


def page_has_no_fare(page: Page) -> bool:
    text = page_body_text(page).lower()
    patterns = [
        r"\bno\s+(?:available\s+)?flights?\b",
        r"\bno\s+(?:available\s+)?fares?\b",
        r"\bunable\s+to\s+find\s+flights?\b",
        r"\bflights?.{0,80}\bnot\s+available\b",
        r"\bfares?.{0,80}\bnot\s+available\b",
        r"\bsold\s+out\b",
    ]
    return any(re.search(p, text) for p in patterns)


def page_has_flights(page: Page) -> bool:
    try:
        return bool(re.search(r"ZL\s?\d{3,4}", page_body_text(page)))
    except Exception:
        return False


def page_has_verification(page: Page) -> bool:
    """
    Detect Rex's human-verification page.
    Uses precise signals — avoids false positives from grecaptcha script tags
    that appear on every Rex page.
    """
    try:
        text = page_body_text(page).lower()
        html = page.content().lower()

        text_markers = [
            "please verify your details",
            "verify you are human",
            "access denied",
            "blocked",
        ]
        if any(m in text for m in text_markers):
            return True

        # .availCont button present = verification page
        if "availcont" in html:
            return True

        # Active reCAPTCHA iframe + txtcaptcha input both present
        if "iframe" in html and "recaptcha" in html and "txtcaptcha" in html:
            return True

        return False
    except Exception:
        return False


def page_is_rex_home(page: Page) -> bool:
    selectors = [
        "label[for*='rbTripType_oneway']",
        "input[id*='rbTripType_oneway']",
        "#ContentPlaceHolder1_BookingHomepageV21_OriginAirport",
        "#datefilter",
    ]
    for sel in selectors:
        try:
            if page.query_selector(sel):
                return True
        except Exception:
            pass
    return False


# ─────────────────────────────────────────────────────────────
#  CAPTCHA / VERIFICATION HANDLING
# ─────────────────────────────────────────────────────────────

def wait_for_brightdata_captcha(page: Page, wait_secs: int = 60):
    """
    Wait for Bright Data Scraping Browser to handle CAPTCHA.

    With Playwright + Scraping Browser we can detect page state directly.
    We exit as soon as the page reaches any meaningful state.
    """
    print(f"   ⏳ Bright Data CAPTCHA solver active — waiting up to {wait_secs}s...")
    deadline = time.time() + wait_secs
    while time.time() < deadline:
        if page_is_rex_home(page):
            break
        if page_has_flights(page):
            break
        if page_has_no_fare(page):
            break
        if page_has_verification(page):
            break  # Verification page — let caller handle reCAPTCHA
        time.sleep(2)
    print("   ✅ CAPTCHA wait complete")


def handle_verification_if_present(page: Page, timeout_secs: int = 90) -> bool:
    """
    Detect Rex's human-verification page and solve CAPTCHA using Bright Data CDP.
    Then click the Continue button (.availCont).
    """
    if not page_has_verification(page):
        return True

    print(f"   🧩 Verification page detected! Triggering Bright Data Captcha.solve (timeout {timeout_secs}s)...")

    # 1. Try solving via Bright Data CDP session
    try:
        client = page.context.new_cdp_session(page)
        res = client.send("Captcha.solve", {"detectTimeout": timeout_secs * 1000})
        status = res.get("status", "unknown")
        print(f"   🧩 CDP: Captcha.solve status: {status}")
        try:
            client.detach()
        except Exception:
            pass
    except Exception as e:
        print(f"   ⚠️  CDP Captcha.solve failed: {e}")
        status = "error"

    # 2. Wait for Continue button (.availCont) to be enabled and click it
    print("   🔒 Waiting for Continue button to become enabled...")
    deadline = time.time() + 15
    while time.time() < deadline:
        try:
            result = page.evaluate("""() => {
                const btn = document.querySelector('.availCont');
                if (!btn) return 'not_found';

                const isDisabled = btn.disabled || btn.hasAttribute('disabled');
                if (!isDisabled) {
                    btn.click();
                    return 'clicked';
                }
                return 'disabled';
            }""")
            if result == 'clicked':
                print("   ✅ Verification Continue button clicked.")
                return True
            elif result == 'not_found':
                print("   ✅ Verification button gone — page moved on.")
                return True
        except Exception:
            pass
        time.sleep(1)

    # 3. Last resort: force enable and click the button
    print("   ⚠️  Continue button still disabled/not clicked — attempting force-click (last resort)...")
    try:
        result = page.evaluate("""() => {
            const btn = document.querySelector('.availCont');
            if (!btn) return 'not_found';
            btn.removeAttribute('disabled');
            btn.disabled = false;
            btn.click();
            return 'force_clicked';
        }""")
        if result in ('force_clicked', 'not_found'):
            print("   ✅ Force-clicked Continue button.")
            return True
    except Exception as e:
        print(f"   ❌ Force-click failed: {e}")

    return False


# ─────────────────────────────────────────────────────────────
#  REX HOMEPAGE NAVIGATION
# ─────────────────────────────────────────────────────────────

def open_rex_home(page: Page, label: str, attempts: int = 3) -> bool:
    """Navigate to Rex homepage and wait for booking form."""
    for attempt in range(1, attempts + 1):
        ts = int(time.time())
        url = f"https://www.rex.com.au/?pw_retry={ts}_{attempt}"
        print(f"   🌐 Rex homepage load (attempt {attempt}/{attempts})...")
        try:
            # Playwright goto — waits for domcontentloaded
            page.goto(url, wait_until="commit", timeout=max(PAGE_TIMEOUT * 1000, 60_000))
        except PWTimeout:
            print(f"   ⚠️  Page load timed out (attempt {attempt}) — checking state anyway...")
        except Exception as e:
            print(f"   ⚠️  Page load exception: {e}")

        # Wait for page to settle
        wait_for_brightdata_captcha(page, wait_secs=15)

        # Handle verification page immediately if present on homepage load
        if page_has_verification(page):
            handle_verification_if_present(page, timeout_secs=CAPTCHA_WAIT)
            time.sleep(3)

        if wait_for_rex_home(page, timeout=45):
            return True

        if attempt < attempts:
            time.sleep(5 * attempt)

    return False


def wait_for_rex_home(page: Page, timeout: int = 90) -> bool:
    """Wait for Rex homepage booking form to become ready."""
    print("   ⏳ Waiting for Rex homepage form...")
    deadline = time.time() + timeout
    while time.time() < deadline:
        if page_is_rex_home(page):
            print("   ✅ Rex booking form ready")
            return True
        # Check for any Continue button on intermediate pages
        # Use evaluate (safer than query_selector_all which can timeout on blank pages)
        try:
            clicked = page.evaluate("""() => {
                const btns = Array.from(document.querySelectorAll('button'));
                for (const btn of btns) {
                    const txt = (btn.innerText || '').trim().toLowerCase();
                    if (txt.includes('continue') && !btn.disabled) {
                        btn.click();
                        return true;
                    }
                }
                return false;
            }""")
            if clicked:
                print("   ✅ Continue button clicked (intermediate page)")
                time.sleep(2)
        except Exception:
            pass
        time.sleep(1.5)
    print("   ❌ Rex homepage form did not appear in time")
    return False


# ─────────────────────────────────────────────────────────────
#  FORM FILL HELPERS
# ─────────────────────────────────────────────────────────────

def click_one_way(page: Page) -> bool:
    """Select the One-way radio button."""
    try:
        result = page.evaluate("""() => {
            const one = document.getElementById('ContentPlaceHolder1_BookingHomepageV21_rbTripType_oneway');
            const ret = document.getElementById('ContentPlaceHolder1_BookingHomepageV21_rbTripType_return');
            if (!one) return false;
            one.checked = true;
            if (ret) ret.checked = false;
            one.dispatchEvent(new Event('change', { bubbles: true }));
            one.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
            return true;
        }""")
        if result:
            print("   ✅ One-way selected")
            return True
    except Exception:
        pass

    # Fallback: try CSS selectors
    for sel in [
        "label[for*='rbTripType_oneway']",
        "input[id*='rbTripType_oneway']",
    ]:
        try:
            el = page.query_selector(sel)
            if el:
                el.click()
                print("   ✅ One-way selected (fallback)")
                return True
        except Exception:
            pass

    print("   ❌ Could not select one-way trip type")
    return False


def select_airport(page: Page, select_id: str, code: str, label: str,
                   wait_for_option: bool = False) -> bool:
    """Select an airport from a <select> dropdown."""
    if wait_for_option:
        # Wait for destination options to load after origin triggers JS update
        deadline = time.time() + 20
        while time.time() < deadline:
            result = page.evaluate("""([id, code]) => {
                const el = document.getElementById(id);
                if (!el) return false;
                return Array.from(el.options).some(opt =>
                    opt.value === code ||
                    (opt.textContent || '').toUpperCase().includes('(' + code + ')')
                );
            }""", [select_id, code])
            if result:
                break
            time.sleep(0.5)

    # Try Playwright's built-in select_option first (cleanest approach)
    try:
        page.select_option(f"#{select_id}", value=code, timeout=5000)
        time.sleep(0.5)
        print(f"   ✅ {label} airport selected: {code}")
        return True
    except Exception:
        pass

    # Fallback: JS direct set
    result = page.evaluate("""([id, code]) => {
        const el = document.getElementById(id);
        if (!el) return { ok: false, reason: 'selector missing' };
        const option = Array.from(el.options).find(opt => {
            const val  = (opt.value || '').trim();
            const text = (opt.textContent || '').trim().toUpperCase();
            return val === code || text.includes('(' + code + ')') || text === code;
        });
        if (!option) return {
            ok: false, reason: 'option missing',
            available: Array.from(el.options).map(o => ({ v: o.value, t: o.textContent.trim() }))
        };
        el.value = option.value;
        option.selected = true;
        el.dispatchEvent(new Event('input',  { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        if (window.jQuery) window.jQuery(el).trigger('change');
        return { ok: true };
    }""", [select_id, code])

    if not result.get("ok"):
        print(f"   ❌ Could not set {label} airport: {code} — {result.get('reason')}")
        avail = result.get("available", [])
        if avail:
            print(f"      Available: {avail[:8]}")
        return False

    time.sleep(0.5)
    print(f"   ✅ {label} airport selected: {code}")
    return True


def set_departure_date(page: Page, target_dt: datetime) -> bool:
    """Set departure date in the Rex booking form."""
    date_text = rex_form_date(target_dt)
    try:
        result = page.evaluate("""(dateText) => {
            const dateInput = document.getElementById('datefilter');
            const dep = document.getElementById('ContentPlaceHolder1_BookingHomepageV21_HDepartureDate');
            const ret = document.getElementById('ContentPlaceHolder1_BookingHomepageV21_HReturnDate');
            if (!dateInput || !dep) return false;
            dateInput.value = dateText;
            dep.value = dateText;
            if (ret) ret.value = '';
            for (const el of [dateInput, dep].filter(Boolean)) {
                el.dispatchEvent(new Event('input',  { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
            }
            return true;
        }""", date_text)
        if result:
            print(f"   📅 Date set: {date_text}")
            return True
    except Exception as e:
        print(f"   ⚠️  Direct date set failed: {e}")

    # Fallback: click calendar and pick the day
    try:
        date_input = page.query_selector("#datefilter")
        if date_input:
            date_input.click()
            time.sleep(1)

            target_month = datetime(target_dt.year, target_dt.month, 1)
            for _ in range(18):
                try:
                    month_els = page.query_selector_all(".daterangepicker .month")
                    if not month_els:
                        break
                    month_text = month_els[0].inner_text().strip()
                    try:
                        shown = datetime.strptime(month_text, "%b %Y")
                    except ValueError:
                        try:
                            shown = datetime.strptime(month_text, "%B %Y")
                        except ValueError:
                            break
                    if shown.year == target_dt.year and shown.month == target_dt.month:
                        break
                    nav = "next" if target_month > shown else "prev"
                    btn = page.query_selector(f".daterangepicker .{nav}")
                    if btn:
                        btn.click()
                    time.sleep(0.4)
                except Exception:
                    break

            day_re = re.compile(rf"^\s*{target_dt.day}\s*$")
            cells = page.query_selector_all("td.available:not(.off)")
            for cell in cells:
                if day_re.match(cell.inner_text()):
                    cell.click()
                    print(f"   📅 Date picked from calendar: {date_text}")
                    return True
    except Exception as e:
        print(f"   ⚠️  Date picker fallback failed: {e}")

    return False


def dismiss_cookie_popup(page: Page):
    """Dismiss cookie/GDPR consent popup if present."""
    try:
        result = page.evaluate("""() => {
            const selectors = [
                '.eupopup-button_1',
                '.eupopup-closebutton',
                '[class*="eupopup"] button',
                '[id*="cookieConsent"] button',
                '[class*="cookie"] button',
            ];
            for (const sel of selectors) {
                const el = document.querySelector(sel);
                if (el) { el.click(); return sel; }
            }
            const overlay = document.querySelector('.eupopup-container, [class*="eupopup"]');
            if (overlay) { overlay.remove(); return 'removed overlay'; }
            return null;
        }""")
        if result:
            print(f"   🍪 Cookie popup dismissed ({result})")
            time.sleep(0.8)
    except Exception:
        pass


def submit_form(page: Page) -> bool:
    """Submit the Rex booking search form."""
    submit_id = "ContentPlaceHolder1_BookingHomepageV21_SubmitBooking"

    dismiss_cookie_popup(page)

    # Playwright: scroll + click directly
    try:
        btn = page.query_selector(f"#{submit_id}")
        if btn:
            btn.scroll_into_view_if_needed()
            time.sleep(0.4)
            btn.click()
            print("   ✅ Form submitted")
            return True
    except Exception as e:
        print(f"   ⚠️  Submit click failed; trying JS: {e}")

    # JS direct click fallback
    try:
        result = page.evaluate("""(id) => {
            const btn = document.getElementById(id);
            if (!btn) return false;
            btn.scrollIntoView({block: 'center'});
            btn.click();
            return true;
        }""", submit_id)
        if result:
            print("   ✅ Form submitted (JS direct click)")
            return True
    except Exception as e:
        print(f"   ❌ JS submit failed: {e}")

    return False


# ─────────────────────────────────────────────────────────────
#  WAIT FOR SEARCH RESULT
# ─────────────────────────────────────────────────────────────

def wait_for_search_result(page: Page, target_dt: datetime, timeout: int = 180) -> tuple[str, str]:
    """
    Wait for Rex search result page to load after form submit.
    Returns: (status, reason)

    Key behaviours:
    - Detects flight results, no-fare, verification page, or timeout
    - Verification handled with smart loop — tracks force-click usage
    - Infinite loop detection via URL tracking
    - Time-budgeted verification waits
    """
    deadline = time.time() + timeout
    verification_strikes = 0
    MAX_VERIFICATION_STRIKES = 6
    force_click_used = False
    last_url_on_verification = ""

    while time.time() < deadline:
        if page_has_flights(page):
            print("   ✅ Flights visible on page")
            return STATUS_SUCCESS, "Flight results loaded"

        if page_has_no_fare(page):
            print("   ℹ️  No fare signal detected")
            return STATUS_NO_FARE, "Rex page indicates no flight/fare available"

        if page_has_verification(page):
            verification_strikes += 1
            current_url = page.url

            # Detect infinite loop — stuck on same URL after force-click
            if force_click_used and current_url == last_url_on_verification:
                print("   ❌ Stuck on same verification page after force-click — giving up")
                return STATUS_BLOCKED, "Verification page loop detected after force-click"

            if verification_strikes > MAX_VERIFICATION_STRIKES:
                print(f"   ❌ Verification page persists after {MAX_VERIFICATION_STRIKES} attempts")
                return STATUS_BLOCKED, f"Verification not cleared after {MAX_VERIFICATION_STRIKES} attempts"

            remaining = max(15, int(deadline - time.time()))
            strikes_left = max(1, MAX_VERIFICATION_STRIKES - verification_strikes + 1)
            wait_budget = min(90, max(15, (remaining - 10) // strikes_left))

            print(f"   🧩 Verification page (strike {verification_strikes}/{MAX_VERIFICATION_STRIKES}) "
                  f"— waiting up to {wait_budget}s for reCAPTCHA...")

            last_url_on_verification = current_url
            clicked = handle_verification_if_present(page, timeout_secs=wait_budget)
            force_click_used = True  # Mark so we don't force-click again

            time.sleep(5 if clicked else 2)
            continue

        # Normal loading — still waiting
        verification_strikes = 0
        time.sleep(2)

    return STATUS_TIMEOUT, f"Timed out after {timeout}s waiting for Rex result"


# ─────────────────────────────────────────────────────────────
#  CORE SCRAPE — ONE DATE, ONE ATTEMPT
# ─────────────────────────────────────────────────────────────

def scrape_one_date(pw_instance, origin: str, dest: str, target_dt: datetime,
                    attempt: int, store: OutputStore) -> tuple[str, list[dict], str]:
    """
    Create fresh Playwright browser context → open Rex → fill form →
    submit → extract → close context.

    Returns: (status, rows, debug_artifacts_str)
    """
    date_str = output_date(target_dt)
    label    = f"{RUN_ID}_{origin}_{dest}_{date_str}_attempt_{attempt}"
    browser  = None
    context  = None
    page     = None

    try:
        browser, context, page = make_page(pw_instance)

        # ── Step 1: Homepage ──────────────────────────────────
        if not open_rex_home(page, f"{label}_home"):
            debug = save_debug(page, f"{label}_home_failed",
                               {"status": "HOME_FAILED", "origin": origin, "dest": dest})
            return STATUS_FAILED, [], debug

        # ── Step 2: One-way ───────────────────────────────────
        if not click_one_way(page):
            debug = save_debug(page, f"{label}_oneway_failed")
            return STATUS_STRUCT, [], debug

        # ── Step 3: Route ─────────────────────────────────────
        origin_id = "ContentPlaceHolder1_BookingHomepageV21_OriginAirport"
        dest_id   = "ContentPlaceHolder1_BookingHomepageV21_DestinationAirport"

        if not select_airport(page, origin_id, origin, "Origin"):
            debug = save_debug(page, f"{label}_origin_failed")
            return STATUS_STRUCT, [], debug

        if not select_airport(page, dest_id, dest, "Destination", wait_for_option=True):
            debug = save_debug(page, f"{label}_dest_failed")
            return STATUS_STRUCT, [], debug

        # ── Step 4: Date ──────────────────────────────────────
        if not set_departure_date(page, target_dt):
            debug = save_debug(page, f"{label}_date_failed")
            return STATUS_STRUCT, [], debug

        # ── Step 5: Submit ────────────────────────────────────
        if not submit_form(page):
            debug = save_debug(page, f"{label}_submit_failed")
            return STATUS_STRUCT, [], debug

        # Wait for page state after submit
        wait_for_brightdata_captcha(page, wait_secs=CAPTCHA_WAIT)

        # Handle verification page immediately after submit
        if page_has_verification(page):
            clicked = handle_verification_if_present(page, timeout_secs=CAPTCHA_WAIT)
            # Give server time to process after the click
            time.sleep(5 if clicked else 2)

        # ── Step 6: Wait for result ───────────────────────────
        status, reason = wait_for_search_result(page, target_dt, timeout=PAGE_TIMEOUT)

        if status == STATUS_NO_FARE:
            row = make_no_fare_row(date_str, origin, dest, reason, attempt - 1)
            return STATUS_NO_FARE, [row], ""

        if status != STATUS_SUCCESS:
            debug = save_debug(page, f"{label}_{status}",
                               {"status": status, "reason": reason})
            return status, [], debug

        # ── Step 7: Extract flights ───────────────────────────
        body_text = page_body_text(page)
        flights   = extract_flights_from_body(body_text, origin, dest, date_str)

        if flights:
            missing = [f for f in flights if f.get("Fare Price") in {"N/A", "-", ""}]
            if missing:
                debug = save_debug(page, f"{label}_missing_price",
                                   {"status": STATUS_STRUCT,
                                    "reason": "Flights found but price missing",
                                    "parsed": len(flights)})
                for f in flights:
                    f["Status"]  = STATUS_STRUCT
                    f["Comment"] = "Price extraction failed — check debug"
                return STATUS_STRUCT, flights, debug
            return STATUS_SUCCESS, flights, ""

        if page_has_no_fare(page):
            row = make_no_fare_row(date_str, origin, dest,
                                   "No fare confirmed after extraction attempt", attempt - 1)
            return STATUS_NO_FARE, [row], ""

        debug = save_debug(page, f"{label}_no_extract",
                           {"status": STATUS_STRUCT,
                            "reason": "Page loaded with ZL but no flights extracted"})
        return STATUS_STRUCT, [], debug

    except Exception as exc:
        comment = f"Unhandled exception: {exc}"
        print(f"   ❌ {comment}")
        traceback.print_exc()
        debug = ""
        if page:
            try:
                debug = save_debug(page, f"{label}_exception",
                                   {"status": STATUS_FAILED, "reason": comment,
                                    "traceback": traceback.format_exc()})
            except Exception:
                pass
        return STATUS_FAILED, [], debug

    finally:
        # Always clean up — context and browser must both be closed
        for obj, name in [(page, "page"), (context, "context"), (browser, "browser")]:
            if obj:
                try:
                    obj.close()
                except Exception:
                    pass
        print("   🔌 Browser context closed")


# ─────────────────────────────────────────────────────────────
#  ROUTE RUNNER
# ─────────────────────────────────────────────────────────────

def run_route(pw_instance, origin: str, dest: str, store: OutputStore):
    dates       = build_date_list()
    origin_name = AIRPORT_MAP.get(origin, origin)
    dest_name   = AIRPORT_MAP.get(dest, dest)

    print(f"\n{'█'*60}")
    print(f"  ROUTE : {origin} ({origin_name}) → {dest} ({dest_name})")
    print(f"  Dates : {dates[0].strftime('%d-%m-%Y')} → {dates[-1].strftime('%d-%m-%Y')}")
    print(f"  Output: {OUTPUT_EXCEL}")
    print(f"  Attempts/date: {MAX_ATTEMPTS}")
    print(f"{'█'*60}\n")

    for idx, target_dt in enumerate(dates, 1):
        date_str = output_date(target_dt)

        print(f"\n{'═'*60}")
        print(f"📅 [{idx}/{len(dates)}]  {target_dt.strftime('%A, %d %b %Y')}")
        print(f"{'─'*60}")

        if RESUME_MODE and store.job_completed(origin, dest, date_str):
            print(f"   ↩️  Already completed (run {RUN_ID}) — skipping")
            continue

        final_status = STATUS_FAILED
        final_rows: list[dict] = []
        final_debug = ""

        for attempt in range(1, MAX_ATTEMPTS + 1):
            if attempt > 1:
                backoff = min(90, 8 * (2 ** (attempt - 2)) + random.uniform(0, 3))
                print(f"   ⏳ Backing off {backoff:.1f}s before attempt {attempt}...")
                time.sleep(backoff)

            print(f"   🔁 Attempt {attempt}/{MAX_ATTEMPTS} — {origin}→{dest} {date_str}")
            status, rows, debug = scrape_one_date(
                pw_instance, origin, dest, target_dt, attempt, store
            )

            final_status = status
            final_rows   = rows
            final_debug  = debug

            if status in COMPLETED_STATUSES:
                break
            if status == STATUS_NO_FARE:
                break

            print(f"   ⚠️  Attempt {attempt} ended: {status}")

        if not final_rows:
            final_rows = [make_failed_row(
                date_str, origin, dest, final_status,
                f"All {MAX_ATTEMPTS} attempts failed",
                MAX_ATTEMPTS - 1,
                final_debug,
            )]

        store.write_rows(origin, dest, date_str, final_rows)

        if final_status == STATUS_SUCCESS:
            print(f"   ✅ {len(final_rows)} flight row(s) saved.")
        elif final_status == STATUS_NO_FARE:
            print(f"   ℹ️  No fare for {date_str}")
        else:
            print(f"   ❌ {final_status}")


# ─────────────────────────────────────────────────────────────
#  CLI ARGS
# ─────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Rex PER↔MJK Playwright Scraper (Bright Data)")
    p.add_argument("--days",               type=int, default=TOTAL_DAYS)
    p.add_argument("--start-offset-days",  type=int, default=0,
                   help="0 = start from today (default). 1 = start from tomorrow.")
    p.add_argument("--output",             default=OUTPUT_EXCEL)
    p.add_argument("--debug-dir",          default=DEBUG_DIR)
    p.add_argument("--log-dir",            default=LOG_DIR)
    p.add_argument("--run-id",             default=RUN_ID)
    p.add_argument("--max-attempts",       type=int, default=MAX_ATTEMPTS)
    p.add_argument("--page-timeout",       type=int, default=PAGE_TIMEOUT)
    p.add_argument("--captcha-wait",       type=int, default=CAPTCHA_WAIT)
    p.add_argument("--resume",             action="store_true",
                   help="Skip dates already completed in the output file.")
    p.add_argument("--routes", default="",
                   help="e.g. PER-MJK,MJK-PER  (leave blank for interactive menu)")
    p.add_argument("--interactive",        action="store_true",
                   help="Run in interactive mode to select routes via terminal menu.")
    return p.parse_args()


def interactive_route_selection() -> list[tuple[str, str]]:
    print("\n" + "═" * 60)
    print("  REX PER ↔ MJK SCRAPER — ROUTE SELECTION")
    print("═" * 60)
    print()
    print("  Available routes:")
    print()
    for i, (o, d) in enumerate(ROUTES, 1):
        print(f"    {i}.  {o} → {d}  ({AIRPORT_MAP.get(o,'?')} → {AIRPORT_MAP.get(d,'?')})")
    print()
    print("    0.  ✅ BOTH ROUTES (PER→MJK + MJK→PER)")
    print()

    while True:
        raw = input("  Enter route number(s) (0 or 1,2 or just 1): ").strip()
        if not raw:
            continue
        tokens = [t.strip() for t in raw.split(",") if t.strip()]
        try:
            selections = [int(t) for t in tokens]
        except ValueError:
            print("  ❌ Please enter numbers only.\n")
            continue

        if 0 in selections:
            print(f"\n  ✅ Both routes selected.\n")
            return list(ROUTES)

        bad = [s for s in selections if s < 1 or s > len(ROUTES)]
        if bad:
            print(f"  ❌ Invalid number(s): {bad}. Must be between 1 and {len(ROUTES)}.\n")
            continue

        chosen = [ROUTES[s - 1] for s in selections]
        print("\n  Selected routes:")
        for o, d in chosen:
            print(f"    ✈️  {o} → {d}  ({AIRPORT_MAP.get(o,'?')} → {AIRPORT_MAP.get(d,'?')})")
        print()
        confirm = input("  Confirm? (y/n): ").strip().lower()
        if confirm in {"y", "yes", ""}:
            return chosen
        print("  Please select again.\n")


# ─────────────────────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    ns = parse_args()

    # Apply CLI config globally
    TOTAL_DAYS   = max(1, ns.days)
    START_OFFSET = max(0, ns.start_offset_days)
    OUTPUT_EXCEL = ns.output
    DEBUG_DIR    = ns.debug_dir
    LOG_DIR      = ns.log_dir
    RUN_ID       = ns.run_id
    MAX_ATTEMPTS = max(1, ns.max_attempts)
    PAGE_TIMEOUT = max(60, ns.page_timeout)
    CAPTCHA_WAIT = max(10, ns.captcha_wait)
    RESUME_MODE  = ns.resume

    # Logging — tee all stdout+stderr to log file as well as console
    log_fh = None
    if LOG_DIR:
        os.makedirs(LOG_DIR, exist_ok=True)
        log_path = os.path.join(LOG_DIR, f"rex_per_mjk_{RUN_ID}_{rex_now():%Y%m%d_%H%M%S}.log")
        log_fh   = open(log_path, "a", encoding="utf-8", errors="replace")

        class _Tee:
            """Write to both original stream and log file."""
            def __init__(self, orig, fh):
                self._orig = orig
                self._fh   = fh
            def write(self, data):
                try:
                    self._orig.write(data)
                except Exception:
                    pass
                try:
                    if not self._fh.closed:
                        self._fh.write(data)
                except Exception:
                    pass
                try:
                    self.flush()
                except Exception:
                    pass
            def flush(self):
                try:
                    self._orig.flush()
                except Exception:
                    pass
                try:
                    if not self._fh.closed:
                        self._fh.flush()
                except Exception:
                    pass
            def isatty(self):
                return False
            # Forward any other attribute lookups to original stream
            def __getattr__(self, name):
                return getattr(self._orig, name)

        sys.stdout = _Tee(sys.stdout, log_fh)
        sys.stderr = _Tee(sys.stderr, log_fh)
        print(f"🧾 Log: {log_path}")

    # Route selection
    if ns.routes.strip():
        routes_to_run = []
        for token in ns.routes.split(","):
            parts = re.split(r"[-:>]", token.strip().upper())
            parts = [p for p in parts if p]
            if len(parts) == 2:
                routes_to_run.append((parts[0], parts[1]))
        if not routes_to_run:
            print("❌ Could not parse --routes; defaulting to BOTH routes.")
            routes_to_run = list(ROUTES)
    elif ns.interactive:
        routes_to_run = interactive_route_selection()
    else:
        # Default to non-interactive mode running both routes
        print("ℹ️  Non-interactive mode — defaulting to BOTH routes (PER-MJK + MJK-PER).")
        routes_to_run = list(ROUTES)

    store = OutputStore(OUTPUT_EXCEL)

    print("\n" + "═" * 60)
    print("  REX PER ↔ MJK SCRAPER — RUN SUMMARY  [Playwright Edition]")
    print("═" * 60)
    for o, d in routes_to_run:
        print(f"  ✈️   {o} → {d}  ({AIRPORT_MAP.get(o,'?')} → {AIRPORT_MAP.get(d,'?')})")
    print(f"  📆 Dates   : {TOTAL_DAYS} days  (from {output_date(build_date_list()[0])}  [today])")
    print(f"  🗂  Output  : {OUTPUT_EXCEL}")
    print(f"  🔁 Run ID  : {RUN_ID}")
    print(f"  ↩️  Resume  : {'Yes — skipping completed dates' if RESUME_MODE else 'No — fresh run'}")
    print(f"  🌐 Engine  : Playwright (Bright Data Scraping Browser CDP)")
    print(f"  🔗 Endpoint: wss://...@{BD_BROWSER_HOST}:{BD_SCRAPING_BROWSER_PORT}")
    print("═" * 60 + "\n")

    route_status = {}

    # Single sync_playwright context wraps the whole run
    with sync_playwright() as pw:
        try:
            for origin, dest in routes_to_run:
                try:
                    run_route(pw, origin, dest, store)
                    route_status[(origin, dest)] = "ok"
                except KeyboardInterrupt:
                    print(f"\n⛔ Stopped at {origin}→{dest}")
                    route_status[(origin, dest)] = "interrupted"
                    break
                except Exception as exc:
                    route_status[(origin, dest)] = f"ERROR: {exc}"
                    print(f"❌ Route {origin}→{dest} fatal error: {exc}")
        finally:
            print("\n" + "═" * 60)
            print("  ROUTE RUN RESULTS")
            print("═" * 60)
            for (o, d) in routes_to_run:
                st = route_status.get((o, d), "not reached")
                if st == "ok":
                    print(f"  ✅  {o} → {d}")
                elif st == "interrupted":
                    print(f"  ⛔  {o} → {d}  (interrupted)")
                else:
                    print(f"  ❌  {o} → {d}  — {st}")
            print("═" * 60)
            print(f"\n📊 Output: {OUTPUT_EXCEL}\n")

            # Send Email Report via Google App Passwords SMTP
            email_subject = f"Rex Scraper Report: Run {RUN_ID}"
            email_body = f"Rex Airlines scraper run completed.\n\n"
            email_body += f"Run ID: {RUN_ID}\n"
            email_body += f"Time Checked: {rex_now().strftime('%Y-%m-%d %H:%M:%S')} (Perth Time)\n"
            email_body += f"Output Excel File: {OUTPUT_EXCEL}\n\n"
            email_body += "Results by Route:\n"
            for (o, d) in routes_to_run:
                st = route_status.get((o, d), "not reached")
                email_body += f"  ✈️  {o} → {d} : {st}\n"

            excel_attachment = OUTPUT_EXCEL if os.path.exists(OUTPUT_EXCEL) else None
            send_email_with_attachment(email_subject, email_body, excel_attachment)

    # Flush, restore original streams, and close log file handle if open
    if log_fh:
        try:
            sys.stdout.flush()
            sys.stderr.flush()
            if hasattr(sys.stdout, "_orig"):
                sys.stdout = sys.stdout._orig
            if hasattr(sys.stderr, "_orig"):
                sys.stderr = sys.stderr._orig
            log_fh.close()
        except Exception:
            pass
